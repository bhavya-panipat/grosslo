"""
salary_revision_export.py — Bulk Salary Revision XLSX export for closing
the Compliance & Savings Audit mode's loop (detects a problem -> generates
the corrected file, doesn't dispatch it).

Deliberately a separate module from app.py's RazorpayX Composite Payout
payload generator (_build_composite_payout). Those are genuinely different
real-world actions: setting up a new hire's payout vs. revising an
existing employee's structure. Merging them would blur two things that
happen at different points in a company's payroll cycle, for different
reasons, through (in RazorpayX Payroll's real product) different flows.

HONESTY LABEL, not confirmed against a live account:
RazorpayX Payroll's real Bulk Salary Revision flow is documented as
template-driven — a "Salary Revision Sheet" with a Default Structure sheet
(org-wide revised CTC) and a Custom Structure sheet (per-employee CTC
breakdown). The exact column headers were not available to check
byte-for-byte against a live template. This module reuses this codebase's
own existing field names (basic, hra, lta, special_allowance, employer_pf,
employer_nps) across the same two-sheet shape, and says so explicitly —
in the API response and inside the file itself (the Read Me sheet) — so
this is never mistaken for a verified, ready-to-upload template.

Zero new tax/compliance logic. Every corrected value here is the
optimizer's own output for the flagged employee, passed in already
computed — this module only formats it into the workbook shape.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

TEMPLATE_HONESTY_LABEL = (
    "Column structure approximated to match RazorpayX Payroll's Bulk Salary "
    "Revision template. Exact headers not verified against a live RazorpayX "
    "account — confirm against the real template before uploading."
)

FIELDS = ["basic", "hra", "lta", "special_allowance", "employer_pf", "employer_nps"]
FIELD_LABELS = {
    "basic": "Basic", "hra": "HRA", "lta": "LTA",
    "special_allowance": "Special Allowance",
    "employer_pf": "Employer PF", "employer_nps": "Employer NPS",
}


def build_salary_revision_workbook(rows: list[dict]) -> Workbook:
    """
    rows: [{"employee_name": str, "ctc": float,
            "current": {basic, hra, lta, special_allowance, employer_pf, employer_nps},
            "corrected": {basic, hra, lta, special_allowance, employer_pf, employer_nps}}, ...]
    Only flagged employees (audit mode's excess-contribution or
    unclaimed-savings rows) should be passed in — this doesn't decide who
    needs revising, it only formats whoever the caller already identified.
    """
    wb = Workbook()

    readme = wb.active
    readme.title = "Read Me"
    readme["A1"] = "grosslo — Bulk Salary Revision export"
    readme["A1"].font = Font(bold=True, size=13)
    readme["A3"] = TEMPLATE_HONESTY_LABEL
    readme["A3"].alignment = Alignment(wrap_text=True)
    readme.column_dimensions["A"].width = 90
    readme["A5"] = "This file was generated from grosslo's Compliance & Savings Audit mode."
    readme["A6"] = "It contains corrected structures for employees already flagged for excess EPFO contribution or unclaimed regime-switch savings — nothing in this file was invented; every corrected value is the deterministic optimizer's own output for that employee."
    readme["A6"].alignment = Alignment(wrap_text=True)
    readme.row_dimensions[6].height = 40
    readme["A8"] = "No live upload to RazorpayX occurs anywhere in this codebase. This file must be reviewed and uploaded manually, the same way any Bulk Salary Revision file would be."
    readme["A8"].alignment = Alignment(wrap_text=True)

    header_fill = PatternFill(start_color="F0ECE4", end_color="F0ECE4", fill_type="solid")
    header_font = Font(bold=True)

    default_sheet = wb.create_sheet("Default Structure")
    default_sheet.append(["Employee Name", "Revised CTC"])
    for cell in default_sheet[1]:
        cell.fill, cell.font = header_fill, header_font
    for row in rows:
        default_sheet.append([row["employee_name"], row["ctc"]])
    default_sheet.column_dimensions["A"].width = 28
    default_sheet.column_dimensions["B"].width = 16

    custom_sheet = wb.create_sheet("Custom Structure")
    custom_header = ["Employee Name", "CTC"] + [FIELD_LABELS[f] for f in FIELDS]
    custom_sheet.append(custom_header)
    for cell in custom_sheet[1]:
        cell.fill, cell.font = header_fill, header_font
    for row in rows:
        corrected = row["corrected"]
        custom_sheet.append([row["employee_name"], row["ctc"]] + [corrected.get(f, 0) for f in FIELDS])
    custom_sheet.column_dimensions["A"].width = 28
    for col in "BCDEFGH":
        custom_sheet.column_dimensions[col].width = 16

    return wb
